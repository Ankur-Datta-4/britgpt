#!/usr/bin/env python3
"""Import national 2x2 / flavor card data from BGPT.xlsx (ignores 'For Sid' sheet)."""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX_CANDIDATES = [
    Path.home() / "Downloads" / "BGPT (1).xlsx",
    Path.home() / "Downloads" / "BGPT.xlsx",
    ROOT / "BGPT.xlsx",
]

# Trendline sheet uses 13 monthly points: May-25 … May-26
TREND_MONTH_LABELS = [
    "May-25",
    "Jun-25",
    "Jul-25",
    "Aug-25",
    "Sep-25",
    "Oct-25",
    "Nov-25",
    "Dec-25",
    "Jan-26",
    "Feb-26",
    "Mar-26",
    "Apr-26",
    "May-26",
]


def resolve_xlsx_path() -> Path:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1]).expanduser()
        if not p.exists():
            raise FileNotFoundError(p)
        return p
    for p in DEFAULT_XLSX_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError(
        "No BGPT xlsx found. Pass path: python scripts/import-bgpt-xlsx.py ~/Downloads/BGPT.xlsx"
    )


def fmt_pct(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        p = v * 100 if abs(v) <= 1 else v
        return f"{p:.1f}%"
    s = str(v).strip()
    if s.endswith("%"):
        return s
    try:
        f = float(s)
        return f"{f * 100:.1f}%" if f <= 1 else f"{f:.1f}%"
    except ValueError:
        return s


def int_idx(v):
    if v is None:
        return None
    return int(round(float(v)))


def import_national_flavors(wb) -> list:
    ws = wb["Data Feeding into 2x2 & Cards"]
    flavors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or str(name).strip() == "Flavor":
            continue
        name = str(name).strip()
        item = {
            "name": name,
            "convVolume": str(row[1]).strip() if row[1] else None,
            "engVolume": str(row[2]).strip() if row[2] else None,
            "convGrowth": fmt_pct(row[3]),
            "engGrowth": fmt_pct(row[4]),
            "diyIndex": int_idx(row[5]),
            "shareabilityIndex": int_idx(row[6]),
            "cravingIndex": int_idx(row[7]),
            "comfortIndex": int_idx(row[8]),
            "curiosityIndex": int_idx(row[9]),
            "giftingIndex": int_idx(row[10]),
            "whyPopular": str(row[11]).strip() if row[11] else None,
            "states": str(row[12]).strip() if row[12] else None,
            "whenTrends": str(row[13]).strip() if row[13] else None,
            "trendType": str(row[14]).strip() if row[14] else None,
            "extensions": str(row[15]).strip() if row[15] else None,
            "brandFit": str(row[16]).strip() if row[16] else None,
        }
        if row[17] is not None:
            try:
                item["fitmentScore"] = round(float(row[17]), 1)
            except (TypeError, ValueError):
                pass
        if row[18]:
            item["brandFitReasoning"] = str(row[18]).strip()
        flavors.append(item)
    return flavors


def import_extension_overrides(wb) -> dict:
    """Sheet: Product Extension Ideas and Exp — curated ideas + reasoning for select flavors."""
    if "Product Extension Ideas and Exp" not in wb.sheetnames:
        return {}
    ws = wb["Product Extension Ideas and Exp"]
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        ideas = str(row[1]).strip() if row[1] else None
        reasoning = str(row[2]).strip() if row[2] else None
        if not ideas and not reasoning:
            continue
        entry = {}
        if ideas:
            entry["extensions"] = ideas
        if reasoning:
            entry["extensionReasoning"] = reasoning
        out[name] = entry
    return out


def merge_extension_overrides(flavors: list, overrides: dict) -> None:
    by_name = {f["name"]: f for f in flavors}
    for name, patch in overrides.items():
        if name in by_name:
            by_name[name].update(patch)
        else:
            print(f"  warn: extension sheet flavor not in national data: {name}")


def import_key_insights(wb) -> dict:
    ws = wb["Key Insights (Select Flavors)"]
    insights = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            insights[str(row[0]).strip()] = str(row[1]).strip()
    return insights


def import_index_definitions(wb) -> dict:
    ws = wb["Variable Definitions"]
    definitions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            definitions[str(row[0]).strip()] = str(row[1]).strip()
    return definitions


def import_trendline(wb) -> dict | None:
    if "Trendline Data" not in wb.sheetnames:
        return None
    ws = wb["Trendline Data"]
    series = {}
    n_months = len(TREND_MONTH_LABELS)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        vals = []
        for i in range(n_months):
            v = row[1 + i] if 1 + i < len(row) else None
            if v is None:
                break
            vals.append(round(float(v), 1))
        if len(vals) == n_months:
            series[name] = vals
        else:
            print(f"  warn: trendline for {name} has {len(vals)} points, expected {n_months}")
    return {"months": TREND_MONTH_LABELS, "series": series}


def main():
    xlsx = resolve_xlsx_path()
    print(f"Reading {xlsx}")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    flavors = import_national_flavors(wb)
    ext_overrides = import_extension_overrides(wb)
    if ext_overrides:
        merge_extension_overrides(flavors, ext_overrides)
        print(f"Merged product extension overrides for {len(ext_overrides)} flavors")

    (ROOT / "lib/demo-national-parsed.json").write_text(
        json.dumps(flavors, indent=2) + "\n", encoding="utf-8"
    )

    insights = import_key_insights(wb)
    (ROOT / "lib/demo-flavor-key-insights.json").write_text(
        json.dumps(insights, indent=2) + "\n", encoding="utf-8"
    )

    definitions = import_index_definitions(wb)
    (ROOT / "lib/demo-flavor-index-definitions.json").write_text(
        json.dumps(definitions, indent=2) + "\n", encoding="utf-8"
    )

    trend = import_trendline(wb)
    if trend:
        (ROOT / "lib/flavor-conv-trend.json").write_text(
            json.dumps(trend, indent=2) + "\n", encoding="utf-8"
        )
        print(f"Wrote trendline for {len(trend['series'])} flavors → lib/flavor-conv-trend.json")

    print(f"Wrote {len(flavors)} flavors → lib/demo-national-parsed.json")
    print(f"Wrote {len(insights)} key insights → lib/demo-flavor-key-insights.json")
    print(f"Wrote {len(definitions)} index definitions → lib/demo-flavor-index-definitions.json")


if __name__ == "__main__":
    main()
